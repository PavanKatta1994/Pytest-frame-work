import openpyxl
from datetime import datetime
with open("data.txt","r") as myfile:
    lines = myfile.readlines()
date = datetime
filedata = []
for line in lines:
    linedata = [line[0:10].strip(), line[10:20].strip(), line[20:30].strip(),
                date.strptime(line[30:].strip(), "%d%m%Y")]
    filedata.append(linedata)
filedata.sort(key=lambda x: x[3], reverse=True)
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "DataParsed"
sheet.cell(1,1).value = "First Name"
sheet.cell(1,2).value = "Middle Name"
sheet.cell(1,3).value = "Surname Name"
sheet.cell(1,4).value = "Date Of Birth"
for i in range(len(filedata)):
    for j in range(len(filedata[i])):
        if filedata[i][j] == "":
            sheet.cell(i+2,j+1).value = "na"
        else:
            if j == 3:
                sheet.cell(i + 2, j + 1).value = datetime.strftime(filedata[i][j], "%d-%B-%y")
            else:
                sheet.cell(i + 2, j + 1).value = filedata[i][j]
workbook.save("ParsedExcel.xlsx")