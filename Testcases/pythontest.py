import openpyxl
import mysql.connector

def ExcelData(filename, sheetname):
    file = "../Testdata/" + filename + ".xlsx"
    excel = openpyxl.load_workbook(file)
    worksheet = excel[sheetname]
    # calculating number of columns as test data
    column_count = 0
    for i in range(worksheet.max_column):
        value = worksheet.cell(1, i + 1).value
        if value != "Test Status" :
            print(value)
            column_count += 1
        else:
            break
    rows = worksheet.max_row
    cols = column_count
    whole_date = []
    for i in range(rows):
        data = []
        for j in range(cols):
            value = worksheet.cell(i + 2, j + 1).value
            data.append(value)
        whole_date.append(data)
    excel.close()
    return whole_date

def update_excel(filename, sheetname, column_name, row_id, value):
    file = "../Testdata/" + filename + ".xlsx"
    excel = openpyxl.load_workbook(file)
    worksheet = excel[sheetname]
    column_count = 1
    for i in range(worksheet.max_column):
        col_val = worksheet.cell(1, i + 1).value
        if col_val == column_name:
            break
        else:
            column_count += 1
    col_id = column_count
    worksheet.cell(row_id+1, column_count).value = value
    excel.save(file)
    excel.close()

def dataquery(schema, query):
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="root",
    )

    cursor = mydb.cursor()
    cursor.execute("use {};".format(schema))
    if ("update" in query.lower()) or ("insert" in query.lower()):
        cursor.execute("{};".format(query))
        mydb.commit()
    else:
        date = []
        cursor.execute("{};".format(query))
        data = cursor.fetchall()
        for d in data:
            print(d)
        mydb.close()
        return data


print(dataquery("pavan","select test_id, user_name, user_password, message, login from login_test"))