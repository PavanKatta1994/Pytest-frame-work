from datetime import datetime
import logging
import inspect
import os
import mysql.connector
import mysql
import softest
from openpyxl import Workbook, load_workbook
from pathlib import Path
import configparser

class Utils(softest.TestCase):

    @staticmethod
    def dataquery(schema, query):
        mydb = mysql.connector.connect(
            host="localhost",
            user="root",
            password="root",
        )

        cursor = mydb.cursor()
        cursor.execute("use {};".format(schema))
        if ("update" in query.lower()) or ("insert" in query.lower()):
            cursor.execute("{};".format(query))
            mydb.commit()
            mydb.close()
        else:
            data = []
            cursor.execute("{};".format(query))
            data = cursor.fetchall()
            for d in data:
                print(d)
            mydb.close()
            return data

    @staticmethod
    def custom_logger(newlog = True):
        config = configparser.ConfigParser()
        config.read("../ConfigFiles/config.ini")
        generate_log = config["Logging"]["GenerateLogs"]
        loglevel = config["Logging"]["LogLevel"]
        if generate_log == "True":
            if newlog:
                try:
                    extn = datetime.now().strftime("%Y-%m-%d_%H%M%S")
                    os.rename("../Logs/Automation_Log.log", "../Logs/Automation_Log_{}.log".format(extn))
                    os.chmod("../Logs/Automation_Log.log", 777)
                except FileNotFoundError:
                    print("No Log file found")
                    print("Proceeding to create new log file")
                else:
                    print("Existing log file renamed")
                    print("New log file to be created")
                # logger_name = inspect.stack()[1][3]
            frame = inspect.currentframe().f_back
            method_name = frame.f_code.co_name
            cls_name = frame.f_locals.get('self', None)
            class_name = cls_name.__class__.__name__ if cls_name else "NoClass"
            logger_name = f"{class_name}.{method_name}"
            logger = logging.getLogger(logger_name)
            logger.setLevel(loglevel)
            if not logger.handlers:
                fh = logging.FileHandler('../Logs/Automation_Log.log', mode='a')
                formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                                              datefmt='%d-%m-%y %I:%M:%S %p')
                fh.setFormatter(formatter)
                logger.addHandler(fh)
            return logger


    @staticmethod
    def get_current_date():
        return datetime.today().date()

    @staticmethod
    def get_day_month_year_from_date(date):
        try:
            date_obj = datetime.strptime(date, "%d/%m/%Y").date()
        except ValueError:
            print("provided Invalid date")
            print("Setting date as today")
            date_obj =  datetime.today().date()
            return date_obj.strftime("%d"),date_obj.strftime("%B"),date_obj.strftime("%Y"),date_obj
        else:
            return date_obj.strftime("%d"),date_obj.strftime("%B"),date_obj.strftime("%Y"),date_obj

    def compare_text_of_elements(self,element_list, value):
        utils = Utils()
        # self.log = utils.custom_logger()
        fail_count = 0
        # print(type(element_list))
        print("no of values to be verified : {}".format(len(element_list)))
        for element in element_list:
            actual_value = element.text
            try:
                assert actual_value == value
                print(f"element {actual_value} was verified : {value}")
            except AssertionError:
                fail_count += 1
        return fail_count

class Excel:

        def __init__(
                self,
                file_name: str,
                sheet_name: str,
                location: str = "../Files",
                delete_existing: bool = False
        ):

            self.file_name = file_name if file_name.endswith(".xlsx") else file_name + ".xlsx"
            self.sheet_name = sheet_name
            self.location = Path(location)
            self.location.mkdir(parents=True, exist_ok=True)

            self.file_path = self.location / self.file_name

            # Delete existing excel if requested
            if delete_existing and self.file_path.exists():
                self.file_path.unlink()
                print(f"Existing excel deleted : {self.file_path}")

            # Create workbook object
            if self.file_path.exists():
                self.workbook = load_workbook(self.file_path)
                print(f"Existing excel loaded : {self.file_path}")
            else:
                self.workbook = Workbook()
                print(f"New excel will be created : {self.file_path}")

            # Create / Load sheet
            if self.sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[self.sheet_name]
                print(f"Using existing sheet : {self.sheet_name}")
            else:
                # Remove default sheet if empty workbook
                if "Sheet" in self.workbook.sheetnames and len(self.workbook.sheetnames) == 1:
                    default_sheet = self.workbook["Sheet"]
                    self.workbook.remove(default_sheet)

                self.sheet = self.workbook.create_sheet(self.sheet_name)
                print(f"New sheet created : {self.sheet_name}")

        # ---------------------------------------------------
        # m1. add_header
        # ---------------------------------------------------
        def add_header(self, columns: list):

            if not columns:
                raise ValueError("Header list cannot be empty")

            for col_num, column_name in enumerate(columns, start=1):
                self.sheet.cell(row=1, column=col_num, value=column_name)

            print("Header added successfully")

        # ---------------------------------------------------
        # m2. add_data
        # ---------------------------------------------------
        def add_data(self, values: list):

            if not values:
                raise ValueError("Data list cannot be empty")

            next_row = self.sheet.max_row + 1

            for col_num, value in enumerate(values, start=1):

                # Convert empty string / None to NULL
                if value == "" or value is None:
                    value = "NULL"

                self.sheet.cell(row=next_row, column=col_num, value=value)

            print(f"Data added successfully at row {next_row}")

        # ---------------------------------------------------
        # Internal helper method
        # ---------------------------------------------------
        def _get_header_map(self):

            headers = {}

            for col in range(1, self.sheet.max_column + 1):
                header_name = self.sheet.cell(row=1, column=col).value

                if header_name:
                    headers[header_name] = col

            return headers

        # ---------------------------------------------------
        # Internal helper method
        # ---------------------------------------------------
        def _find_matching_rows(self, condition: dict):

            headers = self._get_header_map()

            for key in condition.keys():
                if key not in headers:
                    print(f"Column '{key}' not found")
                    return []

            matching_rows = []

            for row in range(2, self.sheet.max_row + 1):

                match = True

                for key, expected_value in condition.items():

                    col_num = headers[key]
                    actual_value = self.sheet.cell(row=row, column=col_num).value

                    if actual_value != expected_value:
                        match = False
                        break

                if match:
                    matching_rows.append(row)

            return matching_rows

        # ---------------------------------------------------
        # m3. modify_column_data
        # ---------------------------------------------------
        def modify_column_data(self, condition: dict, column: str, value):

            headers = self._get_header_map()

            if column not in headers:
                print(f"Column '{column}' not found")
                return

            matching_rows = self._find_matching_rows(condition)

            if not matching_rows:
                print("Condition not satisfied. No matching rows found.")
                return

            target_col = headers[column]

            for row in matching_rows:
                self.sheet.cell(row=row, column=target_col, value=value)

            print(f"{len(matching_rows)} row(s) updated successfully")

        # ---------------------------------------------------
        # m4. read_column_data
        # ---------------------------------------------------
        def read_column_data(self, condition: dict, column: str):

            headers = self._get_header_map()

            if column not in headers:
                print(f"Column '{column}' not found")
                return None

            matching_rows = self._find_matching_rows(condition)

            if not matching_rows:
                return None

            target_col = headers[column]

            result = []

            for row in matching_rows:
                result.append(
                    self.sheet.cell(row=row, column=target_col).value
                )

            if len(result) == 1:
                return result[0]

            return result

        # ---------------------------------------------------
        # m5. save_excel
        # ---------------------------------------------------
        def save_excel(self):

            self.workbook.save(self.file_path)
            print(f"Excel saved successfully : {self.file_path}")